"""Bulk-import worker for the Audit Universe.

A multipart POST to ``/api/auditable-entities/bulk-import/`` creates a
``BulkImportJob`` row and enqueues this task. The task:

  1. Streams the uploaded file via openpyxl (XLSX) or csv (CSV).
  2. Maps each row dict onto the field set accepted by
     ``AuditableEntitySerializer`` — using the same camelCase shape the
     write API expects so the validation rules stay in one place.
  3. Upserts on ``(external_source, external_id)`` when those columns
     are present; otherwise creates a new entity if no name match
     exists, or updates the existing entity by name.
  4. Captures the first 200 row-level errors in ``BulkImportJob.errors``
     for inline UI feedback.

Strict mode aborts the whole transaction on the first failure; lenient
mode wraps each row in a savepoint so a bad row doesn't poison the rest.
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Iterable

from celery import shared_task
from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)


# Field aliases — accept human-friendly column headers from spreadsheet
# templates and map them onto the camelCase keys the serializer expects.
COLUMN_ALIASES: dict[str, str] = {
    "name": "name",
    "entity name": "name",
    "auditable entity": "name",
    "description": "description",
    "department": "departmentName",
    "department / function": "departmentName",
    "department name": "departmentName",
    "department id": "departmentId",
    "business unit": "businessUnitName",
    "business unit id": "businessUnitId",
    "owner": "ownerEmail",
    "primary owner": "ownerEmail",
    "primary owner email": "ownerEmail",
    "secondary owner": "secondaryOwnerEmail",
    "entity type": "entityType",
    "type": "entityType",
    "risk rating": "riskRating",
    "compliance status": "complianceStatus",
    "audit frequency": "auditFrequency",
    "last audit rating": "lastAuditRating",
    "last audit date": "lastAuditDate",
    "next audit date": "nextAuditDate",
    "last audit period": "lastAuditPeriod",
    "primary language": "primaryLanguage",
    "location": "location",
    "headcount": "headcount",
    "operating budget": "operatingBudget",
    "operating budget (usd)": "operatingBudget",
    "estimated man days": "estimatedManDays",
    "estimated man-days": "estimatedManDays",
    "man days": "estimatedManDays",
    "man-days": "estimatedManDays",
    "mandatory to audit": "isMandatoryToAudit",
    "cost center": "costCenterId",
    "cost center id": "costCenterId",
    "tags": "tags",
    "inherent likelihood": "inherentLikelihood",
    "likelihood": "inherentLikelihood",
    "inherent impact": "inherentImpact",
    "impact": "inherentImpact",
    "external source": "external_source",
    "external id": "external_id",
}


def _normalise_header(h: str) -> str:
    return COLUMN_ALIASES.get((h or "").strip().lower(), (h or "").strip())


def _row_to_payload(row: dict, *, lookups: dict) -> tuple[dict, str | None]:
    """Translate a free-form row dict into a serializer-ready payload.

    Returns ``(payload, external_key)``. The ``external_key`` is the
    ``(external_source, external_id)`` tuple to use for idempotent
    upserts, when both columns are present.
    """
    payload: dict = {}
    for raw_key, value in row.items():
        key = _normalise_header(raw_key)
        if not key or value in (None, ""):
            continue
        if key == "departmentName":
            dept = lookups["departments_by_name"].get(str(value).strip().lower())
            if dept:
                payload["departmentId"] = str(dept.id)
        elif key == "businessUnitName":
            bu = lookups["bus_by_name"].get(str(value).strip().lower())
            if bu:
                payload["businessUnitId"] = str(bu.id)
        elif key == "ownerEmail":
            owner = lookups["users_by_email"].get(str(value).strip().lower())
            if owner:
                payload["primaryOwnerId"] = str(owner.pk)
        elif key == "secondaryOwnerEmail":
            owner = lookups["users_by_email"].get(str(value).strip().lower())
            if owner:
                payload["secondaryOwnerId"] = str(owner.pk)
        elif key == "isMandatoryToAudit":
            s = str(value).strip().lower()
            payload[key] = s in ("1", "true", "yes", "y", "x")
        elif key == "tags":
            if isinstance(value, str):
                payload[key] = [t.strip() for t in value.split(",") if t.strip()]
            elif isinstance(value, (list, tuple)):
                payload[key] = [str(t).strip() for t in value if str(t).strip()]
        elif key in ("headcount", "inherentLikelihood", "inherentImpact"):
            try:
                payload[key] = int(float(value))
            except (TypeError, ValueError):
                pass
        elif key in ("operatingBudget", "estimatedManDays"):
            payload[key] = str(value)
        else:
            payload[key] = str(value).strip() if isinstance(value, str) else value

    external_source = payload.pop("external_source", None)
    external_id = payload.pop("external_id", None)
    external_key = (
        f"{external_source}:{external_id}"
        if external_source and external_id
        else None
    )
    if external_source and external_id:
        payload["external_source"] = external_source
        payload["external_id"] = external_id
    return payload, external_key


def _stream_rows(file_field) -> Iterable[dict]:
    """Yield row dicts from a CSV or XLSX file field.

    Selection is by extension; XLSX uses openpyxl in read-only mode.
    """
    name = (file_field.name or "").lower()
    file_field.open("rb")
    try:
        if name.endswith(".xlsx") or name.endswith(".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(filename=file_field, read_only=True, data_only=True)
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            try:
                header = [str(c or "").strip() for c in next(iterator)]
            except StopIteration:
                return
            for raw in iterator:
                if not raw or all(c is None or c == "" for c in raw):
                    continue
                yield dict(zip(header, raw))
        else:
            text = io.TextIOWrapper(file_field, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text)
            for raw in reader:
                yield raw
    finally:
        try:
            file_field.close()
        except Exception:  # noqa: BLE001
            pass


@shared_task(name="iams.audit_universe.process_bulk_import")
def process_bulk_import(job_id: str) -> dict:
    """Drive a ``BulkImportJob`` to completion.

    Synchronously parses → validates → writes. The task is idempotent at
    the job level: re-running a completed job is a no-op (it short-
    circuits on the status check). Within the run, the
    ``external_source/external_id`` pair makes the upserts idempotent so
    a re-imported ERP feed doesn't duplicate rows.
    """
    from iams.domain_serializers import AuditableEntitySerializer
    from iams.models import (
        AuditableEntity,
        BulkImportJob,
        BusinessUnit,
        Department,
    )
    from django.contrib.auth import get_user_model

    User = get_user_model()

    try:
        job = BulkImportJob.objects.get(pk=job_id)
    except BulkImportJob.DoesNotExist:
        logger.info("audit-universe-import: job %s vanished before run", job_id)
        return {"processed": 0, "reason": "missing"}

    if job.status not in (BulkImportJob.STATUS_PENDING, BulkImportJob.STATUS_VALIDATING):
        logger.info("audit-universe-import: job %s already %s", job_id, job.status)
        return {"processed": 0, "reason": "already_processed"}

    lookups = {
        "departments_by_name": {d.name.lower(): d for d in Department.objects.all()},
        "bus_by_name": {b.name.lower(): b for b in BusinessUnit.objects.all()},
        "users_by_email": {u.email.lower(): u for u in User.objects.all() if u.email},
    }

    job.status = BulkImportJob.STATUS_IMPORTING
    job.save(update_fields=["status", "updated_at"])

    created = updated = skipped = total = 0
    errors: list[dict] = []
    strict = job.mode == BulkImportJob.MODE_STRICT

    def run():
        nonlocal created, updated, skipped, total
        for row_index, raw in enumerate(_stream_rows(job.file), start=2):
            # Header is row 1; data starts at row 2.
            total += 1
            payload, _ = _row_to_payload(raw, lookups=lookups)
            if not payload.get("name"):
                skipped += 1
                if len(errors) < 200:
                    errors.append({
                        "row": row_index,
                        "field": "name",
                        "message": "Missing required `name` column.",
                    })
                if strict:
                    raise ValueError("Strict mode: missing required column")
                continue

            # Pick the matching existing row, if any:
            instance = None
            ext_src = payload.get("external_source")
            ext_id = payload.get("external_id")
            if ext_src and ext_id:
                instance = AuditableEntity.all_objects.filter(
                    external_source=ext_src, external_id=ext_id,
                ).first()
            if instance is None:
                instance = AuditableEntity.all_objects.filter(
                    name__iexact=payload["name"],
                ).first()

            if instance is not None:
                payload.setdefault("version", instance.version or 1)

            serializer = AuditableEntitySerializer(
                instance=instance, data=payload, partial=instance is not None,
            )
            try:
                if not serializer.is_valid():
                    skipped += 1
                    for field, msgs in serializer.errors.items():
                        if len(errors) >= 200:
                            break
                        msg = msgs[0] if isinstance(msgs, list) else str(msgs)
                        errors.append({
                            "row": row_index,
                            "field": field,
                            "message": str(msg),
                        })
                    if strict:
                        raise ValueError("Strict mode: bad row")
                    continue

                sid = transaction.savepoint() if not strict else None
                try:
                    obj = serializer.save()
                    if instance is None:
                        created += 1
                    else:
                        updated += 1
                    if sid:
                        transaction.savepoint_commit(sid)
                except Exception as exc:  # noqa: BLE001
                    if sid:
                        transaction.savepoint_rollback(sid)
                    skipped += 1
                    if len(errors) < 200:
                        errors.append({
                            "row": row_index,
                            "field": "_save",
                            "message": str(exc)[:240],
                        })
                    if strict:
                        raise
                else:
                    del obj
            except ValueError:
                # Strict-mode bail-out — already recorded above.
                raise

    try:
        if strict:
            with transaction.atomic():
                run()
        else:
            run()
    except ValueError:
        # Strict-mode failure path — surface the partial counts.
        job.refresh_from_db()
        job.status = BulkImportJob.STATUS_FAILED
        job.total_rows = total
        job.processed = total
        job.created = 0
        job.updated = 0
        job.skipped = total
        job.errors = errors
        job.finished_at = timezone.now()
        job.save()
        try:
            from iams import metrics as m
            m.audit_universe_bulk_imports_total.labels(status=job.status).inc()
        except Exception:  # noqa: BLE001
            pass
        return {"status": job.status, "errors": len(errors)}
    except Exception as exc:  # noqa: BLE001
        logger.exception("audit-universe-import: job %s crashed", job_id)
        job.status = BulkImportJob.STATUS_FAILED
        job.errors = errors + [{"row": 0, "field": "_task", "message": str(exc)[:240]}]
        job.finished_at = timezone.now()
        job.save()
        try:
            from iams import metrics as m
            m.audit_universe_bulk_imports_total.labels(status=job.status).inc()
        except Exception:  # noqa: BLE001
            pass
        return {"status": job.status, "errors": len(job.errors)}

    job.status = (
        BulkImportJob.STATUS_PARTIAL if errors else BulkImportJob.STATUS_COMPLETED
    )
    job.total_rows = total
    job.processed = total
    job.created = created
    job.updated = updated
    job.skipped = skipped
    job.errors = errors
    job.finished_at = timezone.now()
    job.save()

    # Emit Phase-7 Prometheus counters. We tolerate failures here so a
    # broken metrics registry never poisons a successful import.
    try:
        from iams import metrics as m
        m.audit_universe_bulk_imports_total.labels(status=job.status).inc()
        m.audit_universe_bulk_import_rows_total.labels(outcome="created").inc(created)
        m.audit_universe_bulk_import_rows_total.labels(outcome="updated").inc(updated)
        m.audit_universe_bulk_import_rows_total.labels(outcome="skipped").inc(skipped)
    except Exception:  # noqa: BLE001
        logger.exception("metrics: failed to bump audit-universe import counters")

    logger.info(
        "audit-universe-import: job %s finished status=%s created=%d updated=%d skipped=%d",
        job_id, job.status, created, updated, skipped,
    )
    return {"status": job.status, "created": created, "updated": updated, "skipped": skipped}
