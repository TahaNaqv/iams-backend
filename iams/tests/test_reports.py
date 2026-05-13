"""Tests for the Report Generation Engine (Phase 4 Track 2).

Coverage:
  - Each PDF renderer builds the expected Jinja context
  - Each Excel renderer writes the right header + rows
  - render_bytes returns HTML when IAMS_DISABLE_PDF_RENDER=1
  - The Celery task creates the file and flips status → completed
  - generate-report endpoint creates a ReportJob and runs the renderer eagerly
  - Download endpoint returns 409 while pending, 404 when failed, URL on completed
  - Excel exports require export_reports permission
  - Job listing is scoped to caller unless manage_settings
  - Audit log captures report_job_created
  - Unknown kind returns 400 with the supported list
"""
from __future__ import annotations

import os
from datetime import date, timedelta
from decimal import Decimal

import pytest

from iams.models import (
    Audit,
    AuditLogEntry,
    AuditableEntity,
    CorrectiveAction,
    EntityRiskScore,
    Finding,
    ReportJob,
    RiskFactor,
    RiskFactorWeight,
    RiskScoringModel,
)
from iams.reports import (
    AnnualPlanRenderer,
    AuditSummaryRenderer,
    CAPsExcelRenderer,
    CAPStatusRenderer,
    FindingsExcelRenderer,
    FindingTrendsRenderer,
    RENDERERS,
)

pytestmark = pytest.mark.django_db


# ──────────────────────────────────────────────────────────────────────
# Helpers / fixtures
# ──────────────────────────────────────────────────────────────────────
@pytest.fixture(autouse=True)
def disable_pdf_render(monkeypatch):
    """The dev/CI environment doesn't have pango/cairo. Render to HTML bytes."""
    monkeypatch.setenv("IAMS_DISABLE_PDF_RENDER", "1")


@pytest.fixture
def audit():
    return Audit.objects.create(
        title="Q1 Treasury", department="Finance", lead_auditor="L",
        status="In Progress", priority="High", risk_rating="High",
        start_date=date(2026, 1, 1), end_date=date(2026, 3, 31),
        scope="Treasury operations", objectives="Validate controls",
        completion_percent=60, findings_count=2,
    )


@pytest.fixture
def findings(audit):
    f1 = Finding.objects.create(
        audit=audit, title="Wire SoD gap", department="Finance",
        severity="Critical", status="Open", owner="o@iams.test",
        due_date=date(2026, 4, 15),
        description="d", root_cause="rc", recommendation="r",
        created_date=date(2026, 2, 10),
    )
    f2 = Finding.objects.create(
        audit=audit, title="Recon timing", department="Finance",
        severity="Medium", status="In Progress", owner="o@iams.test",
        due_date=date(2026, 5, 1),
        description="d", root_cause="rc", recommendation="r",
        created_date=date(2026, 2, 15),
    )
    return [f1, f2]


@pytest.fixture
def caps(findings):
    return [
        CorrectiveAction.objects.create(
            finding=findings[0], title="Deploy dual approval",
            owner="treasury@iams.test", due_date=date(2026, 5, 30),
            status="In Progress", priority="High",
            description="d", progress=40, department="Finance",
        ),
        CorrectiveAction.objects.create(
            finding=findings[1], title="Late CAP",
            owner="treasury@iams.test", due_date=date.today() - timedelta(days=5),
            status="Open", priority="Medium",
            description="d", progress=10, department="Finance",
        ),
    ]


# ══════════════════════════════════════════════════════════════════════
# Renderer contexts
# ══════════════════════════════════════════════════════════════════════
def test_audit_summary_context_includes_severity_counts(audit, findings):
    ctx = AuditSummaryRenderer().gather_context({"audit_id": str(audit.id)})
    assert ctx["audit"].pk == audit.pk
    assert ctx["severity_counts"]["Critical"] == 1
    assert ctx["severity_counts"]["Medium"] == 1
    assert ctx["severity_counts"]["High"] == 0
    assert ctx["severity_counts"]["Low"] == 0
    assert "Q1 Treasury" in ctx["report_title"]


def test_audit_summary_missing_audit_id_raises():
    from iams.reports.base import RendererError
    with pytest.raises(RendererError, match="audit_id"):
        AuditSummaryRenderer().gather_context({})


def test_finding_trends_filters_by_period_year(audit, findings):
    ctx = FindingTrendsRenderer().gather_context({"period": "2026"})
    assert ctx["total"] == 2
    by_severity = {row["severity"]: row["count"] for row in ctx["by_severity"]}
    assert by_severity["Critical"] == 1


def test_finding_trends_filters_by_quarter(audit, findings):
    ctx = FindingTrendsRenderer().gather_context({"period": "2026-Q1"})
    assert ctx["total"] == 2  # both created Feb 2026


def test_cap_status_overdue_includes_days_late(audit, findings, caps):
    ctx = CAPStatusRenderer().gather_context({})
    assert ctx["total"] == 2
    # The "Late CAP" has due_date 5 days ago
    assert ctx["overdue_count"] >= 1
    late = next(c for c in ctx["overdue"] if c.title == "Late CAP")
    assert late.days_late == 5


def test_annual_plan_requires_year_and_model():
    from iams.reports.base import RendererError
    with pytest.raises(RendererError, match="year and scoring_model_id"):
        AnnualPlanRenderer().gather_context({})


def test_annual_plan_renders_top_entities(super_admin):
    factor = RiskFactor.objects.create(code="impact", name="Impact", scale_min=1, scale_max=5)
    factor2 = RiskFactor.objects.create(code="likelihood", name="Likelihood", scale_min=1, scale_max=5)
    model = RiskScoringModel.objects.create(
        name="Default", version="1.0",
        formula=RiskScoringModel.FORMULA_WEIGHTED_SUM,
        is_active=True, high_risk_threshold=Decimal("60"),
    )
    RiskFactorWeight.objects.create(scoring_model=model, factor=factor, weight=Decimal("1"))
    RiskFactorWeight.objects.create(scoring_model=model, factor=factor2, weight=Decimal("1"))

    e1 = AuditableEntity.objects.create(name="Top", department="X", owner="o", risk_rating="Medium", status="Active")
    from iams.risk_engine import record_score
    record_score(e1, model=model, factor_values={"impact": 5, "likelihood": 5}, by_user=super_admin)

    ctx = AnnualPlanRenderer().gather_context({"year": 2026, "scoring_model_id": str(model.id)})
    assert ctx["year"] == 2026
    assert len(ctx["scores"]) == 1
    assert ctx["scores"][0].entity.name == "Top"


# ══════════════════════════════════════════════════════════════════════
# Excel renderers
# ══════════════════════════════════════════════════════════════════════
def test_findings_excel_writes_header_and_rows(audit, findings):
    payload = FindingsExcelRenderer().render_bytes({})
    assert len(payload) > 100  # not empty
    # Parse it back
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws.title == "Findings"
    header = [c.value for c in ws[1]]
    assert "Severity" in header
    assert "Status" in header
    # 2 findings + header
    assert ws.max_row == 3


def test_findings_excel_filter_by_severity(audit, findings):
    payload = FindingsExcelRenderer().render_bytes({"severity": "Critical"})
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws.max_row == 2  # header + 1


def test_caps_excel_includes_days_late_for_overdue(audit, findings, caps):
    payload = CAPsExcelRenderer().render_bytes({})
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    by_title = {r[1]: r for r in rows}
    assert by_title["Late CAP"][8] == 5  # days late column


# ══════════════════════════════════════════════════════════════════════
# Registry + render_bytes shape
# ══════════════════════════════════════════════════════════════════════
def test_registry_contains_all_expected_kinds():
    expected = {
        ReportJob.KIND_AUDIT_SUMMARY,
        ReportJob.KIND_FINDING_TRENDS,
        ReportJob.KIND_CAP_STATUS,
        ReportJob.KIND_ANNUAL_PLAN,
        ReportJob.KIND_FINDINGS_EXCEL,
        ReportJob.KIND_CAPS_EXCEL,
        ReportJob.KIND_TIME_ENTRIES_EXCEL,
    }
    assert expected.issubset(RENDERERS.keys())


def test_render_bytes_returns_html_when_pdf_disabled(audit, findings):
    payload = AuditSummaryRenderer().render_bytes({"audit_id": str(audit.id)})
    assert b"<html" in payload.lower()
    assert b"Q1 Treasury" in payload or "Q1 Treasury".encode() in payload


# ══════════════════════════════════════════════════════════════════════
# Job lifecycle: renderer.run + Celery task
# ══════════════════════════════════════════════════════════════════════
def test_run_completes_job_and_writes_file(audit, findings, super_admin):
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY, title="Test",
        parameters={"audit_id": str(audit.id)},
        requested_by=super_admin,
    )
    renderer = AuditSummaryRenderer()
    renderer.run(job)
    job.refresh_from_db()
    assert job.status == ReportJob.STATUS_COMPLETED
    assert job.output_file
    assert job.file_size_kb > 0
    assert job.error == ""


def test_run_failed_job_records_error(super_admin):
    """Missing audit_id → RendererError → status=failed."""
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY, parameters={},
        requested_by=super_admin,
    )
    AuditSummaryRenderer().run(job)
    job.refresh_from_db()
    assert job.status == ReportJob.STATUS_FAILED
    assert "audit_id" in job.error


def test_celery_task_dispatches_to_right_renderer(audit, findings, super_admin):
    from iams.tasks import generate_report
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY,
        parameters={"audit_id": str(audit.id)},
        requested_by=super_admin,
    )
    result = generate_report(str(job.pk))
    assert result["rendered"] is True
    job.refresh_from_db()
    assert job.status == ReportJob.STATUS_COMPLETED


def test_celery_task_unknown_kind_fails_gracefully(super_admin):
    from iams.tasks import generate_report
    # Bypass model choice validation by setting kind via update()
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY, parameters={"audit_id": "0"*32},
        requested_by=super_admin,
    )
    ReportJob.objects.filter(pk=job.pk).update(kind="nonexistent_kind")
    result = generate_report(str(job.pk))
    assert result["rendered"] is False
    job.refresh_from_db()
    assert job.status == ReportJob.STATUS_FAILED
    assert "No renderer" in job.error


def test_celery_task_notifies_requester_on_completion(audit, findings, super_admin):
    from iams.models import Notification
    from iams.tasks import generate_report
    Notification.objects.filter(recipient=super_admin, module="Reports").delete()
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY,
        parameters={"audit_id": str(audit.id)},
        requested_by=super_admin,
    )
    generate_report(str(job.pk))
    notif = Notification.objects.filter(recipient=super_admin, module="Reports").first()
    assert notif is not None
    assert "Report ready" in notif.title


# ══════════════════════════════════════════════════════════════════════
# API endpoints
# ══════════════════════════════════════════════════════════════════════
def test_api_generate_report_creates_job_and_runs_eagerly(
    authed_client, super_admin, audit, findings
):
    AuditLogEntry.objects.filter(action="export").delete()
    res = authed_client(super_admin).post(
        "/api/reports/generate/",
        {
            "kind": ReportJob.KIND_AUDIT_SUMMARY,
            "title": "Q1 Audit",
            "parameters": {"audit_id": str(audit.id)},
        },
        format="json",
    )
    assert res.status_code == 201, res.content
    body = res.json()
    # Eager Celery → already completed by the time the response returns
    job = ReportJob.objects.get(pk=body["id"])
    assert job.status == ReportJob.STATUS_COMPLETED
    assert AuditLogEntry.objects.filter(details__event="report_job_created").exists()


def test_api_generate_report_unknown_kind_400(authed_client, super_admin):
    res = authed_client(super_admin).post(
        "/api/reports/generate/",
        {"kind": "bogus", "parameters": {}},
        format="json",
    )
    assert res.status_code == 400
    body = res.json()
    assert "supportedKinds" in body


def test_api_excel_requires_export_reports_permission(
    authed_client, audit, findings,
):
    """``Department Head`` lacks ``export_reports`` — POST for Excel kinds → 403."""
    from django.contrib.auth import get_user_model
    from iams.models import Permission, Role, UserProfile
    User = get_user_model()
    role = Role.objects.create(name="VR-only", is_super_admin=False)
    p_view = Permission.objects.create(key="view_reports", name="view_reports", module="reports")
    role.permissions.set([p_view])
    user = User.objects.create_user(
        username="vr", email="vr@iams.test", password="TestPassword123!",
    )
    UserProfile.objects.create(user=user, role=role, department="X", status="Active")

    client = authed_client(user)
    res = client.post(
        "/api/reports/generate/",
        {"kind": ReportJob.KIND_FINDINGS_EXCEL, "parameters": {}},
        format="json",
    )
    assert res.status_code == 403


def test_api_download_409_while_pending(authed_client, super_admin):
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY,
        parameters={"audit_id": "00000000-0000-0000-0000-000000000000"},
        requested_by=super_admin,
        status=ReportJob.STATUS_PENDING,
    )
    res = authed_client(super_admin).get(f"/api/reports/jobs/{job.id}/download/")
    assert res.status_code == 409


def test_api_download_404_when_failed(authed_client, super_admin):
    job = ReportJob.objects.create(
        kind=ReportJob.KIND_AUDIT_SUMMARY, parameters={},
        requested_by=super_admin, status=ReportJob.STATUS_FAILED,
        error="audit_id is required.",
    )
    res = authed_client(super_admin).get(f"/api/reports/jobs/{job.id}/download/")
    assert res.status_code == 404
    body = res.json()
    assert "error" in body


def test_api_download_returns_url_when_completed(
    authed_client, super_admin, audit, findings
):
    res = authed_client(super_admin).post(
        "/api/reports/generate/",
        {"kind": ReportJob.KIND_AUDIT_SUMMARY,
         "parameters": {"audit_id": str(audit.id)}},
        format="json",
    )
    job_id = res.json()["id"]
    res2 = authed_client(super_admin).get(f"/api/reports/jobs/{job_id}/download/")
    assert res2.status_code == 200
    assert "url" in res2.json()


def test_api_list_jobs_scoped_to_caller(
    authed_client, audit_manager, auditor_user, audit, findings
):
    """audit_manager doesn't have manage_settings — only sees own jobs."""
    # Manager creates a job
    res = authed_client(audit_manager).post(
        "/api/reports/generate/",
        {"kind": ReportJob.KIND_AUDIT_SUMMARY,
         "parameters": {"audit_id": str(audit.id)}},
        format="json",
    )
    own_id = res.json()["id"]
    # Auditor creates one too
    res2 = authed_client(auditor_user).post(
        "/api/reports/generate/",
        {"kind": ReportJob.KIND_AUDIT_SUMMARY,
         "parameters": {"audit_id": str(audit.id)}},
        format="json",
    )
    other_id = res2.json()["id"]

    body = authed_client(audit_manager).get("/api/reports/jobs/").json()
    rows = body["results"] if isinstance(body, dict) else body
    ids = {r["id"] for r in rows}
    assert own_id in ids
    assert other_id not in ids


def test_api_list_jobs_admin_sees_all(
    authed_client, super_admin, audit_manager, audit, findings
):
    # Manager creates a job
    res = authed_client(audit_manager).post(
        "/api/reports/generate/",
        {"kind": ReportJob.KIND_AUDIT_SUMMARY,
         "parameters": {"audit_id": str(audit.id)}},
        format="json",
    )
    other_id = res.json()["id"]
    body = authed_client(super_admin).get("/api/reports/jobs/").json()
    rows = body["results"] if isinstance(body, dict) else body
    ids = {r["id"] for r in rows}
    assert other_id in ids


def test_api_rbac_view_reports_required(authed_client, db, roles):
    """Users without view_reports get 403 on /reports/jobs/ and /reports/generate/."""
    from django.contrib.auth import get_user_model
    from iams.models import Permission, Role, UserProfile
    User = get_user_model()
    role = Role.objects.create(name="No-VR", is_super_admin=False)
    role.permissions.set([])
    user = User.objects.create_user(
        username="nvr", email="nvr@iams.test", password="TestPassword123!",
    )
    UserProfile.objects.create(user=user, role=role, department="X", status="Active")
    client = authed_client(user)
    assert client.get("/api/reports/jobs/").status_code == 403
    assert client.post("/api/reports/generate/", {"kind": "audit_summary"}, format="json").status_code == 403
